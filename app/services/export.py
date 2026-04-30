from io import BytesIO
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.models import MemberApplication, Payment
from app.services.labels import APPLICATION_TYPE_LABELS, APPLICANT_MODE_LABELS, AUTO_CHECK_LABELS, FEE_LABELS, ROLE_LABELS, label


def _header(ws, labels: list[str]) -> None:
    ws.append(labels)
    fill = PatternFill("solid", fgColor="111111")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def _autosize(ws) -> None:
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)


def build_members_export(db: Session) -> bytes:
    rows = db.scalars(
        select(MemberApplication)
        .options(joinedload(MemberApplication.payment))
        .order_by(MemberApplication.created_at.desc())
    ).all()

    wb = Workbook()
    ws_members = wb.active
    ws_members.title = "Участники"
    _header(
        ws_members,
        [
            "ID",
            "ФИО члена",
            "Тип заявителя",
            "Фамилия заявителя",
            "Имя заявителя",
            "Отчество заявителя",
            "Гражданство",
            "Фамилия члена",
            "Имя члена",
            "Отчество члена",
            "Категория по бланку",
            "Иное: пояснение",
            "С Уставом ознакомлен(а)",
            "Telegram ID",
            "Username",
            "Мобильный телефон",
            "Домашний телефон",
            "Email",
            "Область",
            "Город",
            "Улица",
            "Дом",
            "Квартира",
            "Дата рождения",
            "Паспорт №",
            "Паспорт выдан",
            "Дата заявления",
            "Место работы / учебы",
            "Мать / законный представитель",
            "Мать: место работы, должность",
            "Отец / законный представитель",
            "Отец: место работы, должность",
            "Тип заявления",
            "Год",
            "Статус заявки",
            "Создано",
        ],
    )

    ws_payments = wb.create_sheet("Оплаты")
    _header(
        ws_payments,
        [
            "ID заявки",
            "ФИО члена",
            "Тип взноса",
            "Назначение оплаты",
            "Ожидаемая сумма",
            "Указанная сумма",
            "Дата оплаты",
            "Плательщик",
            "Номер операции",
            "Канал оплаты",
            "Автопроверка",
            "Решение бота",
            "Статус админа",
            "Комментарий",
            "Кто проверил",
            "Когда проверил",
        ],
    )

    ws_checks = wb.create_sheet("Проверки")
    _header(ws_checks, ["ID заявки", "ФИО члена", "Файл", "SHA256", "Размер", "MIME", "Заметки проверки"])

    for app in rows:
        payment: Payment | None = app.payment
        ws_members.append(
            [
                app.id,
                app.full_name,
                label(APPLICANT_MODE_LABELS, app.applicant_mode),
                app.applicant_last_name,
                app.applicant_first_name,
                app.applicant_middle_name,
                app.citizenship,
                app.member_last_name,
                app.member_first_name,
                app.member_middle_name,
                label(ROLE_LABELS, app.role),
                app.role_other,
                "да" if app.statute_accepted else "нет",
                app.telegram_id,
                app.telegram_username,
                app.phone_mobile or app.phone,
                app.phone_home,
                app.email,
                app.region,
                app.city,
                app.street,
                app.house,
                app.apartment,
                app.birth_date.isoformat() if app.birth_date else "",
                app.passport_number,
                app.passport_issued_by,
                app.statement_date.isoformat() if app.statement_date else "",
                app.workplace,
                app.mother_full_name,
                app.mother_workplace_position,
                app.father_full_name,
                app.father_workplace_position,
                label(APPLICATION_TYPE_LABELS, app.application_type),
                app.membership_year,
                app.status,
                app.created_at.isoformat() if app.created_at else "",
            ]
        )

        if payment:
            ws_payments.append(
                [
                    app.id,
                    app.full_name,
                    payment.fee_type,
                    label(FEE_LABELS, payment.fee_type),
                    float(payment.expected_amount),
                    float(payment.paid_amount) if payment.paid_amount is not None else "",
                    payment.payment_date.isoformat() if payment.payment_date else "",
                    payment.payer_full_name,
                    payment.operation_id,
                    payment.payment_channel,
                    payment.auto_check_status,
                    label(AUTO_CHECK_LABELS, payment.auto_check_status),
                    payment.admin_status,
                    payment.admin_comment,
                    payment.reviewed_by_telegram_id,
                    payment.reviewed_at.isoformat() if payment.reviewed_at else "",
                ]
            )
            try:
                notes = "; ".join(json.loads(payment.auto_check_notes or "[]"))
            except json.JSONDecodeError:
                notes = payment.auto_check_notes
            ws_checks.append(
                [
                    app.id,
                    app.full_name,
                    payment.receipt_original_name,
                    payment.receipt_sha256,
                    payment.receipt_size,
                    payment.receipt_content_type,
                    notes,
                ]
            )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        _autosize(ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
